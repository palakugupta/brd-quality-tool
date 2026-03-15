"""
BRD Quality Assessment – Excel Report Generator
"""

import sqlite3
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BRAND_DARK = "1E2A3A"
WHITE = "FFFFFF"
RED_BG = "FDE8E8"
YELLOW_BG = "FFF8E1"
GREEN_BG = "E8F5E9"

SEVERITY_COLOURS = {
    "critical": RED_BG,
    "major": YELLOW_BG,
    "minor": GREEN_BG,
}

# ── FIXED: error_type values must match exactly what detectors insert ──
ERROR_TYPES = [
    "different_data",
    "incomplete_data",
    "hallucination",
    "depth_mismatch",
    "duplicate_data",
    "platform_constraint",           # was platform_constraints
    "process_flow_error",            # was process_flow_violation
    "missing_process_step",          # was missing_process_steps
    "business_rule_violation",
    "role_responsibility_violation",
    "organization_mismatch",
    "process_dependency_violation",
    "terminology_drift",
]

ERROR_LABELS = {
    "different_data":                "Different Data",
    "incomplete_data":               "Incomplete Data",
    "hallucination":                 "Hallucination",
    "depth_mismatch":                "Depth Mismatch",
    "duplicate_data":                "Duplicate Data",
    "platform_constraint":           "Platform Constraint",
    "process_flow_error":            "Process Flow Issue",
    "missing_process_step":          "Missing Process Step",
    "business_rule_violation":       "Business Rule Violation",
    "role_responsibility_violation": "Role Responsibility Violation",
    "organization_mismatch":         "Organization Mismatch",
    "process_dependency_violation":  "Process Dependency Issue",
    "terminology_drift":             "Terminology Drift",
}


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
HEADER_FONT = Font(name="Arial", bold=True, color=WHITE)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color=WHITE)


def _fetch_findings(conn, brd_doc_id: int):

    sql = """
        SELECT f.finding_id, f.chunk_id, f.error_type, f.severity, f.line_number,
               f.description, f.source_reference, f.rule_id, f.detected_timestamp,
               c.start_line, c.end_line
        FROM findings f
        LEFT JOIN chunks c ON f.chunk_id = c.chunk_id
        WHERE c.doc_id = ?
        ORDER BY f.line_number
    """

    cur = conn.execute(sql, (brd_doc_id,))
    rows = cur.fetchall()
    cols = [c[0] for c in cur.description]
    findings = [dict(zip(cols, r)) for r in rows]

    # deduplicate
    unique = {}
    for f in findings:
        key = (f["line_number"], f["error_type"], f["description"])
        if key not in unique:
            unique[key] = f

    return list(unique.values())


def _fetch_coverage(conn, brd_doc_id: int):

    row = conn.execute(
        "SELECT COALESCE(line_count, 0) FROM documents WHERE doc_id = ?",
        (brd_doc_id,),
    ).fetchone()
    total = row[0] if row else 0

    # ── FIXED: scope flagged count to current BRD via chunk join ──
    flagged = conn.execute(
        """
        SELECT COUNT(DISTINCT f.line_number)
        FROM findings f
        JOIN chunks c ON f.chunk_id = c.chunk_id
        WHERE f.line_number IS NOT NULL
          AND c.doc_id = ?
        """,
        (brd_doc_id,),
    ).fetchone()[0]

    coverage = ((total - flagged) / total * 100) if total else 0

    return total, flagged, coverage


def _fetch_latest_brd_doc_id(conn) -> int:
    row = conn.execute(
        "SELECT doc_id FROM documents WHERE doc_type='output_brd' ORDER BY doc_id DESC LIMIT 1"
    ).fetchone()
    return row[0] if row else None


def _fetch_run_meta(conn):

    def latest(doc_type):
        row = conn.execute(
            "SELECT filename FROM documents WHERE doc_type=? ORDER BY doc_id DESC LIMIT 1",
            (doc_type,),
        ).fetchone()
        return row[0] if row else "N/A"

    return {
        "sow": latest("input_sow"),
        "mom": latest("input_mom"),
        "brd": latest("output_brd"),
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _build_summary(ws, findings, meta, total_lines, flagged_lines, coverage):

    ws.title = "Summary"
    ws.merge_cells("A1:F1")

    title = ws["A1"]
    title.value = "BRD Quality Assessment Report"
    title.font = TITLE_FONT
    title.fill = _fill(BRAND_DARK)
    title.alignment = CENTER

    row = 3

    meta_rows = [
        ("SOW File",        meta["sow"]),
        ("MoM File",        meta["mom"]),
        ("BRD File",        meta["brd"]),
        ("Generated On",    meta["generated"]),
        ("Total BRD Lines", total_lines),
        ("Lines Flagged",   flagged_lines),
    ]

    for label, val in meta_rows:
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, val)
        row += 1

    row += 2

    headers = ["Error Type", "Total", "Critical", "Major", "Minor"]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row, i, h)
        cell.font = HEADER_FONT
        cell.fill = _fill(BRAND_DARK)
        cell.alignment = CENTER
        cell.border = _border()

    summary = {}
    for f in findings:
        et = f["error_type"]
        sv = f["severity"]
        summary.setdefault(et, {"critical": 0, "major": 0, "minor": 0})
        if sv in summary[et]:
            summary[et][sv] += 1

    row += 1

    for etype in ERROR_TYPES:
        counts = summary.get(etype, {"critical": 0, "major": 0, "minor": 0})
        total = sum(counts.values())
        ws.cell(row, 1, ERROR_LABELS.get(etype, etype))
        ws.cell(row, 2, total)
        ws.cell(row, 3, counts["critical"])
        ws.cell(row, 4, counts["major"])
        ws.cell(row, 5, counts["minor"])
        row += 1


DETAIL_HEADERS = [
    "Finding ID",
    "Error Type",
    "Line #",
    "Chunk Range",
    "Severity",
    "Description",
    "Source Reference",
]


def _build_all_sheet(ws, findings):

    ws.title = "All Findings"
    ws.append(DETAIL_HEADERS)

    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = _fill(BRAND_DARK)
        c.alignment = CENTER
        c.border = _border()

    for f in findings:
        start = f.get("start_line")
        end = f.get("end_line")
        chunk_range = f"{start}-{end}" if start and end else ""

        row = [
            f.get("finding_id"),
            ERROR_LABELS.get(f.get("error_type"), f.get("error_type")),
            f.get("line_number"),
            chunk_range,
            f.get("severity"),
            f.get("description"),
            f.get("source_reference"),
        ]

        ws.append(row)

        r = ws.max_row
        sev = (f.get("severity") or "").lower()
        bg = SEVERITY_COLOURS.get(sev)

        for c in ws[r]:
            if bg:
                c.fill = _fill(bg)
            c.border = _border()
            c.alignment = LEFT
            c.font = BODY_FONT


def generate_excel_report(db_path: str) -> bytes:

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    brd_doc_id = _fetch_latest_brd_doc_id(conn)

    findings = _fetch_findings(conn, brd_doc_id)
    total, flagged, coverage = _fetch_coverage(conn, brd_doc_id)
    meta = _fetch_run_meta(conn)

    wb = Workbook()
    ws = wb.active

    _build_summary(ws, findings, meta, total, flagged, coverage)

    ws = wb.create_sheet()
    _build_all_sheet(ws, findings)

    conn.close()

    buffer = io.BytesIO()
    wb.save(buffer)

    return buffer.getvalue()